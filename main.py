import pathlib
import openpyxl
import logging
import json
import logging.config
import time
import sys
from scrape_amazon_for_isbn_options import scrape_amazon_for_isbn_options
from exceptions import ServiceUnavailableError

def chunk(full_list, size_of_chunk):
    for i in range(0, len(full_list),size_of_chunk):
        yield full_list[i:i+size_of_chunk]

def main():
    #-----------------------------Set up logging-------------------------#
    if not pathlib.Path.cwd().joinpath('logs').exists():
        pathlib.Path.cwd().joinpath('logs').mkdir()
    if not pathlib.Path.cwd().joinpath('file-dump').exists():
        pathlib.Path.cwd().joinpath('file-dump').mkdir()
 
    with open('logging_config.json','r',encoding='utf-8') as fObject:
        log_config = json.load(fObject)
    logging.config.dictConfig(log_config)

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)

    isbn_book = openpyxl.load_workbook("Live_ISBNs_201908.xlsx",read_only=True,data_only=True)
    isbn_sheet = isbn_book.active

    all_isbns = [row[1].value for row in isbn_sheet.iter_rows(min_row=2)]

    CHUNK_SIZE = 50 
    
    if not pathlib.Path("start_index_config.json").exists():
        start_index_config = {'start_index':0}
        with open("start_index_config.json","w") as fObject:
            json.dump(start_index_config,fObject)


    with open("start_index_config.json","r") as fObject:
        start_index_config = json.load(fObject)
    START_INDEX = start_index_config['start_index']

    logger.debug(f"START_INDEX = {START_INDEX}")

    for isbns in chunk(all_isbns[START_INDEX:], CHUNK_SIZE):
        logger.debug(f"Number of isbns: {len(isbns)}")
        try:
            item_results = scrape_amazon_for_isbn_options(isbns)
        except ServiceUnavailableError as e:
            last_isbn_processed = e.args[0]
            index_of_isbn = all_isbns.index(last_isbn_processed)
            start_index_config['start_index'] = index_of_isbn
            with open("start_index_config.json","w") as fObject:
                json.dump(start_index_config, fObject)
            logger.error(f"Service Unavailable.Saved index of {last_isbn_processed} to start_index_config.json")
            item_results = e.args[2]
            write_item_results_to_excel(item_results)
            break

        write_item_results_to_excel(item_results)



def write_item_results_to_excel(item_results):
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)

    output_book = openpyxl.Workbook()
    output_sheet = output_book.active
    output_sheet.title = "ISBNs_and_Seller_Data"

    output_sheet.append(["ISBN13",
                     "Seller Name", 
                     "Link to seller's amazon page",
                     "Condition of product (New/Used)",
                     "Shipped from",
                     "Price"])

    date_time = time.strftime("%Y%m%d_%H%M%S")
    logger.debug(f"Writing to Book ISBNs_and_Seller_Data_{date_time}.xlsx")
    for isbn,results in item_results.items():
        for olpOffer in results:
            output_sheet.append([isbn,
                             olpOffer.olpSeller,
                             olpOffer.olpSellerLink,
                             olpOffer.olpCondition,
                             olpOffer.olpDelivery,
                             olpOffer.olpPrice])

    try:
        output_book.save(f"file-dump/ISBNs_and_Seller_Data_{date_time}.xlsx")
        logger.debug(f"File ISBNS_and_Seller_Data_{date_time}.xlsx written successfully")
    except Exception as e:
        logger.error("Could not save file ISBNS_and_Seller_Data_{date_time}.xlsx")


        
if __name__ == '__main__':
    main()






